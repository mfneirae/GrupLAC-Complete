#
#
# #############################################################################
#       Copyright (c) 2018 Universidad Nacional de Colombia All Rights Reserved.
#
#             This work was made as a development to improve data collection
#       for self-assessment and accreditation processes in the Vicedeanship
#       of academic affairs in the Engineering Faculty of the Universidad
#       Nacional de Colombia and is licensed under a Creative Commons
#       Attribution-NonCommercial - ShareAlike 4.0 International License
#       and MIT Licence.
#
#       by Manuel Embus.
#
#       For more information write me to jai@mfneirae.com
#       Or visit my webpage at https://mfneirae.com/
# #############################################################################
#
#
import openpyxl, sys, os, time, logging
global COD_PRODUCTO
start_time = time.time()
Dir = os.getcwd()
os.chdir(Dir+"/Bin")
sys.path.append('../Bin/')
#List of /bin files
import init
import datosbasicos
import prodbibliografica
import prodtecnica
import apropiacion
import obras
import actividades
#end of /bin files
os.chdir(Dir)
condition = 0;
while condition != 1:
    try:
        print ("------> Seleccione la forma en la que desea obtener la información:")
        print ("1) Imprimir datos en CSV y en Insert")
        print ("2) Imprimir datos en Insert para MySQL")
        print ("3) Imprimir datos en CSV")
        mode = int(input('-> Seleccione una opción: '))
        if mode == 1 or mode == 2 or mode == 3:
            condition = 1
        else:
            print ("El varlor escogido no es valido")
    except ValueError:
        print ("Not a number")

wb = openpyxl.load_workbook('./Input/Base.xlsx')
# wb = openpyxl.load_workbook('./Input/Base.xlsx')
sheet = wb['Sheet1']
total = sheet.max_row +1
COD_PRODUCTO = 1;
init.inicio()
LOG_FILENAME = './Logs/Registros.log'
logging.basicConfig(filename=LOG_FILENAME,level=logging.DEBUG,
                        format = '%(asctime)s:%(levelname)s:%(message)s')
LEVELS = {'debug': logging.DEBUG,
          'info': logging.INFO,
          'warning': logging.WARNING,
          'error': logging.ERROR,
          'critical': logging.CRITICAL}
if len(sys.argv) > 1:
    level_name = sys.argv[1]
    level = LEVELS.get(level_name, logging.NOTSET)
    logging.basicConfig(level=level)
print("------> Los GrupLAC han sido cargados, Estado: " + str(1/(total-1)*100) + "%")
for q in range(2,total):
    coduapa = sheet['A'+str(q)].value
    codhermes = sheet['B'+str(q)].value
    codcolciencias = sheet['C'+str(q)].value
    nombregi = sheet['D'+str(q)].value
    dnilider = sheet['E'+str(q)].value
    my_url = sheet['F'+str(q)].value
    datosbasicos.datosextract()
    datosbasicos.institucionesextract()
    datosbasicos.lineasextract()
    datosbasicos.sectoresextract()
    datosbasicos.integrantesextract()
    prodbibliografica.articulosextract()
    from prodbibliografica import contarticulo
    COD_PRODUCTO = int("".join(str(x) for x in contarticulo))
    prodbibliografica.librosextract()
    from prodbibliografica import contlibros
    COD_PRODUCTO = int("".join(str(x) for x in contlibros))
    prodbibliografica.caplibrosextract()
    from prodbibliografica import contcaplibros
    COD_PRODUCTO = int("".join(str(x) for x in contcaplibros))
    prodbibliografica.doctraextract()
    from prodbibliografica import contdoctra
    COD_PRODUCTO = int("".join(str(x) for x in contdoctra))
    prodbibliografica.otrapubdivextract()
    from prodbibliografica import contotrapubdiv
    COD_PRODUCTO = int("".join(str(x) for x in contotrapubdiv))
    prodbibliografica.otrosarticulosextract()
    from prodbibliografica import contotrosarticulos
    COD_PRODUCTO = int("".join(str(x) for x in contotrosarticulos))
    prodbibliografica.otroslibrosextract()
    from prodbibliografica import contotroslibros
    COD_PRODUCTO = int("".join(str(x) for x in contotroslibros))
    prodbibliografica.traduccionesextract()
    from prodbibliografica import conttraducciones
    COD_PRODUCTO = int("".join(str(x) for x in conttraducciones))
    prodtecnica.cartasextract()
    from prodtecnica import contcartas
    COD_PRODUCTO = int("".join(str(x) for x in contcartas))
    prodtecnica.consultoriasextract()
    from prodtecnica import contconsultorias
    COD_PRODUCTO = int("".join(str(x) for x in contconsultorias))
    prodtecnica.disenosiextract()
    from prodtecnica import contdisenosi
    COD_PRODUCTO = int("".join(str(x) for x in contdisenosi))
    prodtecnica.esquemas_trazadosextract()
    from prodtecnica import contesquemas_trazados
    COD_PRODUCTO = int("".join(str(x) for x in contesquemas_trazados))
    prodtecnica.innovaciones_en_gestionextract()
    from prodtecnica import continnovaciones_en_gestion
    COD_PRODUCTO = int("".join(str(x) for x in continnovaciones_en_gestion))
    prodtecnica.innovaciones_procedimientosextract()
    from prodtecnica import continnovaciones_procedimientos
    COD_PRODUCTO = int("".join(str(x) for x in continnovaciones_procedimientos))
    prodtecnica.variedad_animalextract()
    from prodtecnica import contvariedad_animal
    COD_PRODUCTO = int("".join(str(x) for x in contvariedad_animal))
    prodtecnica.variedad_vegetalextract()
    from prodtecnica import contvariedad_vegetal
    COD_PRODUCTO = int("".join(str(x) for x in contvariedad_vegetal))
    prodtecnica.planta_pilotoextract()
    from prodtecnica import contplanta_piloto
    COD_PRODUCTO = int("".join(str(x) for x in contplanta_piloto))
    prodtecnica.otros_productos_tecnicosextract()
    from prodtecnica import contotros_productos_tecnicos
    COD_PRODUCTO = int("".join(str(x) for x in contotros_productos_tecnicos))
    prodtecnica.regulaciones_normasextract()
    from prodtecnica import contregulaciones_normas
    COD_PRODUCTO = int("".join(str(x) for x in contregulaciones_normas))
    prodtecnica.guias_clinicasextract()
    from prodtecnica import contguias_clinicas
    COD_PRODUCTO = int("".join(str(x) for x in contguias_clinicas))
    prodtecnica.prototiposextract()
    from prodtecnica import contprototipos
    COD_PRODUCTO = int("".join(str(x) for x in contprototipos))
    prodtecnica.reglamentos_tecnicosextract()
    from prodtecnica import contreglamentos_tecnicos
    COD_PRODUCTO = int("".join(str(x) for x in contreglamentos_tecnicos))
    prodtecnica.signos_distintivosextract()
    from prodtecnica import contsignos_distintivos
    COD_PRODUCTO = int("".join(str(x) for x in contsignos_distintivos))
    prodtecnica.software_registradoextract()
    from prodtecnica import contsoftware_registrado
    COD_PRODUCTO = int("".join(str(x) for x in contsoftware_registrado))
    prodtecnica.proyectos_leyextract()
    from prodtecnica import contproyectos_ley
    COD_PRODUCTO = int("".join(str(x) for x in contproyectos_ley))
    prodtecnica.empresas_base_tecextract()
    from prodtecnica import contempresas_base_tec
    COD_PRODUCTO = int("".join(str(x) for x in contempresas_base_tec))
    apropiacion.ediciones_apropiacionextract()
    from apropiacion import contediciones_apropiacion
    COD_PRODUCTO = int("".join(str(x) for x in contediciones_apropiacion))
    apropiacion.eventos_cientificosextract()
    from apropiacion import conteventos_cientificos
    COD_PRODUCTO = int("".join(str(x) for x in conteventos_cientificos))
    apropiacion.informes_investigacionextract()
    from apropiacion import continformes_investigacion
    COD_PRODUCTO = int("".join(str(x) for x in continformes_investigacion))
    apropiacion.redes_conocimientosextract()
    from apropiacion import contredes_conocimientos
    COD_PRODUCTO = int("".join(str(x) for x in contredes_conocimientos))
    apropiacion.contenidos_impresosextract()
    from apropiacion import contcontenidos_impresos
    COD_PRODUCTO = int("".join(str(x) for x in contcontenidos_impresos))
    apropiacion.contenidos_multimediaextract()
    from apropiacion import contcontenidos_multimedia
    COD_PRODUCTO = int("".join(str(x) for x in contcontenidos_multimedia))
    apropiacion.contenido_virtualextract()
    from apropiacion import contcontenido_virtual
    COD_PRODUCTO = int("".join(str(x) for x in contcontenido_virtual))
    apropiacion.estrategias_comunicacionextract()
    from apropiacion import contestrategias_comunicacion
    COD_PRODUCTO = int("".join(str(x) for x in contestrategias_comunicacion))
    apropiacion.estrategias_pedagogicasextract()
    from apropiacion import contestrategias_pedagogicas
    COD_PRODUCTO = int("".join(str(x) for x in contestrategias_pedagogicas))
    apropiacion.participacion_ciudadana_proyectosextract()
    from apropiacion import contparticipacion_ciudadana_proyectos
    COD_PRODUCTO = int("".join(str(x) for x in contparticipacion_ciudadana_proyectos))
    apropiacion.participacion_ciudadana_espaciosextract()
    from apropiacion import contparticipacion_ciudadana_espacios
    COD_PRODUCTO = int("".join(str(x) for x in contparticipacion_ciudadana_espacios))
    obras.obras_productosextract()
    from obras import contobras_productos
    COD_PRODUCTO = int("".join(str(x) for x in contobras_productos))
    actividades.asesorias_programaextract()
    from actividades import contasesorias_programa
    COD_PRODUCTO = int("".join(str(x) for x in contasesorias_programa))
    actividades.cursos_corta_duracionextract()
    from actividades import contcursos_corta_duracion
    COD_PRODUCTO = int("".join(str(x) for x in contcursos_corta_duracion))
    actividades.trabajos_dirigidosextract()
    from actividades import conttrabajos_dirigidos
    COD_PRODUCTO = int("".join(str(x) for x in conttrabajos_dirigidos))
    actividades.jurado_comisionesextract()
    from actividades import contjurado_comisiones
    COD_PRODUCTO = int("".join(str(x) for x in contjurado_comisiones))
    actividades.comites_evaluacionextract()
    from actividades import contcomites_evaluacion
    COD_PRODUCTO = int("".join(str(x) for x in contcomites_evaluacion))
    actividades.demas_trabajosextract()
    from actividades import contdemas_trabajos
    COD_PRODUCTO = int("".join(str(x) for x in contdemas_trabajos))
    actividades.proyectos_grupoextract()
    from actividades import contproyectos_grupo
    COD_PRODUCTO = int("".join(str(x) for x in contproyectos_grupo))
    print("El grupo: " + nombregi + " ha sido  procesado, Estado: " + str(q/(total-1)*100) + "%")
    # from datosbasicos import conteventos
    # COD_PRODUCTO = int("".join(str(x) for x in conteventos))
    if q==total-1:
        logging.shutdown()
        print ("------> Escribiendo las bases de datos.")
        if mode == 1:
            import printcsv
            import printinsert
        elif mode == 2:
            import printinsert
        else:
             import printcsv
        print ("-----------------------------------------------------------------------------------------------")
        print ("")
        print ("------> ¡Extracción Exitosa!")
        print ("------> La información se encuentra en la carpeta: Resultados.")
        print ("------> Tiempo de ejecución: %s Minutos." % ((time.time() - start_time)/60))
        print ("")
        print ("***********************************************************************************************")
        sys.exit()
    COD_PRODUCTO = 1
